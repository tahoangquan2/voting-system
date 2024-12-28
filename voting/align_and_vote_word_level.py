import csv
from collections import Counter, defaultdict
import pandas as pd
import re

import unicodedata


# Expanded mapping of accents to their corresponding numbers
ACCENT_TO_NUMBER = {
    "á": 1, "à": 2, "ã": 3, "ả": 4, "ạ": 5,
    "é": 1, "è": 2, "ẽ": 3, "ẻ": 4, "ẹ": 5,
    "í": 1, "ì": 2, "ĩ": 3, "ỉ": 4, "ị": 5,
    "ó": 1, "ò": 2, "õ": 3, "ỏ": 4, "ọ": 5,
    "ú": 1, "ù": 2, "ũ": 3, "ủ": 4, "ụ": 5,
    "ý": 1, "ỳ": 2, "ỹ": 3, "ỷ": 4, "ỵ": 5,
    "ấ": 1, "ầ": 2, "ẫ": 3, "ẩ": 4, "ậ": 5,
    "ế": 1, "ề": 2, "ễ": 3, "ể": 4, "ệ": 5,
    "ố": 1, "ồ": 2, "ỗ": 3, "ổ": 4, "ộ": 5,
    "ắ": 1, "ằ": 2, "ẵ": 3, "ẳ": 4, "ặ": 5,
    "ớ": 1, "ờ": 2, "ỡ": 3, "ở": 4, "ợ": 5,
    "ứ": 1, "ừ": 2, "ữ": 3, "ử": 4, "ự": 5#,
    #"đ": 9  # For 'đ', assigning a unique number
}

# Mapping of accented characters to their base (unmarked) characters
ACCENT_TO_BASE = {
    "á": "a", "à": "a", "ã": "a", "ả": "a", "ạ": "a",
    "é": "e", "è": "e", "ẽ": "e", "ẻ": "e", "ẹ": "e",
    "í": "i", "ì": "i", "ĩ": "i", "ỉ": "i", "ị": "i",
    "ó": "o", "ò": "o", "õ": "o", "ỏ": "o", "ọ": "o",
    "ú": "u", "ù": "u", "ũ": "u", "ủ": "u", "ụ": "u",
    "ý": "y", "ỳ": "y", "ỹ": "y", "ỷ": "y", "ỵ": "y",
    "â": "â", "ấ": "â", "ầ": "â", "ẫ": "â", "ẩ": "â", "ậ": "â",
    "ê": "ê", "ế": "ê", "ề": "ê", "ễ": "ê", "ể": "ê", "ệ": "ê",
    "ô": "o", "ố": "o", "ồ": "o", "ỗ": "o", "ổ": "o", "ộ": "o",
    "ă": "ă", "ắ": "ă", "ằ": "ă", "ẵ": "ă", "ẳ": "ă", "ặ": "ă",
    "ơ": "ơ", "ớ": "ơ", "ờ": "ơ", "ỡ": "ơ", "ở": "ơ", "ợ": "ơ",
    "ư": "ư", "ứ": "ư", "ừ": "ư", "ữ": "ư", "ử": "ư", "ự": "ư"
}



def normalize_vietnamese(word):
    normalized_word = ""
    accent_number = ""
    word = word.lower()
    for char in word:
        if char in ACCENT_TO_NUMBER:
            normalized_word += ACCENT_TO_BASE[char]  # Replace with the base character
            accent_number = str(ACCENT_TO_NUMBER[char])  # Set the corresponding accent number
        else:
            normalized_word += char  # Add the character as is if no accent

    return normalized_word + accent_number




import difflib

def load_words(file_path):
    """Load words from the given text file."""
    with open(file_path, 'r', encoding='utf-8') as file:
        return [line.strip() for line in file]

def load_normalize_words(file_path):
    """Load words from the given text file."""
    with open(file_path, 'r', encoding='utf-8') as file:
        return [normalize_vietnamese(line.strip()) for line in file]

def find_similar_words(word, word_list, max_results=5):
    """
    Find similar words using difflib.
    - Returns up to `max_results` similar words from the word list.
    """
    similar_list  = []
    normalize_word = normalize_vietnamese(word)
    cnt = 0
    for check in word_list:
        normalize_check = normalize_vietnamese(check)
        similar, dis = words_similar(normalize_word, normalize_check, threshold= 0.15)

        if similar:
            similar_list.append(check)
            cnt += 1

        if(cnt >= max_results):
            break
    return similar_list



def check_word_in_file(word, vietnam_word_list, normalize_vietnam_list):
    """
    Check if a word is in the file and return suggestions if it's not found.
    - If found, return "Yes".
    - If not, return "No" and a list of similar words.
    """

    normalize_word = normalize_vietnamese(word)
    normalize_word = re.sub(r'[.,:_;?/"()…]', '', normalize_word)
    if normalize_word in normalize_vietnam_list:
        similar_words = [word]
        return True, similar_words
    else:
        similar_words = find_similar_words(word, vietnam_word_list)
        return False, similar_words


def normalize_text(text):
    # Remove newline characters
    text = re.sub(r'\s*[\n\r]+\s*', ' ', text)

    # Convert "~" and "=" to "-"
    text = text.replace("~", "-").replace("=", "-")

    # Convert "«»" and to '""'
    text = text.replace("« ", '"')
    text = text.replace(" »", '"')

    text = text.replace("_", '')
    text = text.replace("~", '')
    text = text.replace("—", '-')

    # Convert '...' to '…'
    text = re.sub(r'\.\.\.', '…', text)

    # Ensure punctuation is followed by a space
    text = re.sub(r'([….,;:?!])\s', r'\1 ', text)

    # Remove spaces before punctuation
    text = re.sub(r'\s+([….,!?;:])', r'\1', text)

    # Ensure only one space between words
    text = re.sub(r'\s+', ' ', text)

    # Ensure a space follows "-" at the start of the text
    text = re.sub(r'^-\s*', '- ', text)

    # Strip leading and trailing spaces
    text = text.strip()

    # Return normalized text
    return text

def normalize_word(word):
    """
    Normalize a word to lowercase and remove accents (diacritical marks).

    Args:
        word (str): The word to normalize.

    Returns:
        str: The normalized word.
    """

    if(word ==  'None'):
        return word
    # Convert to lowercase
    word = word.lower()

    # Remove accents by decomposing the Unicode characters and filtering
    word = ''.join(
        char for char in unicodedata.normalize('NFD', word)
        if unicodedata.category(char) != 'Mn'
    )
    word = re.sub(r'-', '', word)
    word = re.sub(r'f', 't', word)
    word = re.sub(r'j', 'i', word)


    return word


def levenshtein_distance(a, b):
    """Calculate the Levenshtein distance between two strings."""
    len_a, len_b = len(a), len(b)
    dp = [[0] * (len_b + 1) for _ in range(len_a + 1)]

    # Initialize base cases
    for i in range(len_a + 1):
        dp[i][0] = i  # Cost of deleting all characters from `a`
    for j in range(len_b + 1):
        dp[0][j] = j  # Cost of inserting all characters into `a`

    # Compute distances
    for i in range(1, len_a + 1):
        for j in range(1, len_b + 1):
            if a[i - 1] == b[j - 1]:
                cost = 0  # No cost if characters match
            else:
                cost = 1  # Substitution cost
            dp[i][j] = min(
                dp[i - 1][j] + 1,      # Deletion
                dp[i][j - 1] + 1,      # Insertion
                dp[i - 1][j - 1] + cost  # Substitution
            )

    return dp[len_a][len_b]


def words_similar(word1, word2, threshold=0.5):
    """
    Check if two words are similar based on Levenshtein distance
    after normalization.

    Args:
        word1 (str): First word.
        word2 (str): Second word.
        threshold (float): Maximum allowed distance as a fraction of word length.

    Returns:
        bool: True if the words are similar enough, False otherwise.
    """
    if(word1 == 'None' and word1 == word2):
        return True, 4
    if(word1 == 'None'):
        return False, len(word2)
    if(word2 == 'None'):
        return False, len(word1)
    # Normalize both words
    word1 = normalize_word(word1)
    word2 = normalize_word(word2)
    # Calculate Levenshtein distance
    distance = levenshtein_distance(word1, word2)
    if(distance <= threshold * max(len(word2), len(word1))):
      return True , distance
    else:
      return False , distance

def MED_to_word(sen1, sen2):
    # Initialize cache for MED calculation and operations tracking
    sen1 = sen1.strip()
    sen2 = sen2.strip()
    seq1 = sen1.split(' ')
    seq2 = sen2.split(' ')
    cache = [[float("inf")] * (len(seq2) + 1) for _ in range(len(seq1) + 1)]
    ops = [[None] * (len(seq2) + 1) for _ in range(len(seq1) + 1)]

    # Fill base cases
    for j in range(len(seq2) + 1):
        cache[len(seq1)][j] = len(seq2) - j
        ops[len(seq1)][j] = 'insert'  # Need to insert remaining qn chars
    for i in range(len(seq1) + 1):
        cache[i][len(seq2)] = len(seq1) - i
        ops[i][len(seq2)] = 'delete'  # Need to delete remaining seq1 chars

    # Fill the cache and ops table
    for i in range(len(seq1) - 1, -1, -1):
        for j in range(len(seq2) - 1, -1, -1):
            compare, dis = words_similar(seq1[i],seq2[j])
            if compare:
                cache[i][j] = cache[i + 1][j + 1]
                ops[i][j] = 'match'  # Characters match, move diagonally
            else:
                # Consider all operations: insert, delete, substitute
                insert_cost = 1 + cache[i][j + 1]
                delete_cost = 1 + cache[i + 1][j]
                substitute_cost = 1 + cache[i + 1][j + 1]

                # Choose the operation with the minimum cost
                if insert_cost <= delete_cost and insert_cost <= substitute_cost:
                    cache[i][j] = insert_cost
                    ops[i][j] = 'insert'
                elif delete_cost <= insert_cost and delete_cost <= substitute_cost:
                    cache[i][j] = delete_cost
                    ops[i][j] = 'delete'
                else:
                    cache[i][j] = substitute_cost
                    ops[i][j] = 'substitute'

    # Backtrack
    aligned_seq1, aligned_seq2 = [], []
    i, j = 0, 0
    while i < len(seq1) or j < len(seq2):
        if i < len(seq1) and j < len(seq2) and ops[i][j] == 'match':
            aligned_seq1.append(seq1[i])
            aligned_seq2.append(seq2[j])
            i += 1
            j += 1
        elif i < len(seq1) and ops[i][j] == 'delete':
            aligned_seq1.append(seq1[i])
            aligned_seq2.append('None')
            i += 1
        elif j < len(seq2) and ops[i][j] == 'insert':
            aligned_seq1.append('None')
            aligned_seq2.append(seq2[j])
            j += 1
        elif i < len(seq1) and j < len(seq2) and ops[i][j] == 'substitute':
            aligned_seq1.append(seq1[i])
            aligned_seq2.append(seq2[j])
            i += 1
            j += 1

    aligned_seq1 = ' '.join(aligned_seq1)
    aligned_seq2 = ' '.join(aligned_seq2)
    return aligned_seq1, aligned_seq2


def align_multiple_sequences(sequences, best_ocr = 1):
    cnt = len(sequences)
    try_case = 0
    while try_case < len(sequences):
        aligned_sequences = sequences[:]
        limit = 0
        base_sequence = aligned_sequences.pop(try_case)
        while limit < 10:
            # Align each remaining sequence to the base sequence
            updated_sequences = []
            for seq in aligned_sequences:
                aligned_seq, base_sequence = MED_to_word(seq, base_sequence)
                updated_sequences.append(aligned_seq)

            # Update the list of aligned sequences
            aligned_sequences = updated_sequences

            # Check if all sequences are aligned to the same length


            if all(len(seq. split(' ')) == len(base_sequence.split(' ')) for seq in aligned_sequences):
                aligned_sequences.append(base_sequence)
                return aligned_sequences, 0



            limit += 1

        try_case += 1


    # If can't align return the best ocr result
    res = []
    for i in range(cnt):
        res.append(sequences[best_ocr])
    return res, 1


def read_ocr_inputs(csv_file):
    ocr_outputs = []
    with open(csv_file, mode='r', encoding='utf-8') as file:
        reader = csv.reader(file)
        for row in reader:
            # Skip the filename column and extract the next 4 columns
            ocr_outputs.append(row[1:5])  # Columns 2 to 5 (0-based index)

    return ocr_outputs


import string

def most_common_end_punctuation(strings):
    """
    Find the most common ending punctuation in a list of strings.
    :param strings: List of strings
    :return: The most common ending punctuation and its count
    """
    end_punctuations = []

    # Define punctuation characters
    punctuation_set = {'?', '!', '.', ';' , ',', ']', '}', '%' , '$', '…'}

    # Collect ending punctuation from each string
    for s in strings:
        if s and s[-1] in punctuation_set:
            end_punctuations.append(s[-1])

    # Use Counter to find the most common ending punctuation
    if end_punctuations:
        counter = Counter(end_punctuations)
        return counter.most_common(1)[0][0]  # Return the most common punctuation and its count
    else:
        return '' # No ending punctuation found

def most_common_start_punctuation(strings):
    """
    Find the most common start punctuation in a list of strings.
    :param strings: List of strings
    :return: The most common start punctuation and its count
    """
    start_punctuations = []

    # Define punctuation characters
    punctuation_set = ['-', '"', '(' , '{', '[']

    # Collect ending punctuation from each string
    for s in strings:
        if s and s[0] in punctuation_set:
            start_punctuations.append(s[0])

    # Use Counter to find the most common ending punctuation
    if start_punctuations:
        counter = Counter(start_punctuations)
        return counter.most_common(1)[0][0]  # Return the most common punctuation and its count
    else:
        return '' # No ending punctuation found

def is_integer(s):
    """
    Check if a string represents an integer.
    :param s: Input string
    :return: True if the string represents an integer, False otherwise
    """
    try:
        int(s)  # Try converting the string to an integer
        return True
    except ValueError:
        return False

def vote(list_word, vietnam_word_list, normalize_vietnam_list):

    vote_container = []
    weight_container = []
    begin_punctuation = most_common_start_punctuation(list_word)
    end_punctuation = most_common_end_punctuation(list_word)
    for i in range(len(list_word)):
        nw = re.sub(r'[.,:_;?/"()…]', '', list_word[i])

        in_list, similar_words = check_word_in_file(nw, vietnam_word_list, normalize_vietnam_list)
        if in_list:
            for j in range(len(list_word)):
                if(i != j):
                    now = re.sub(r'[.,:_;?/"()…]', '', list_word[j])
                    if(now == nw):
                        return list_word[i], 2004

    for word in list_word:
        org_word = word
        word = re.sub(r'[.,:_;?/"()…]', '', word)
        if word == 'None':
            continue
        if is_integer(word):
            vote_container.append(word)
            weight_container.append(1)
            continue
        in_list, similar_words = check_word_in_file(word, vietnam_word_list, normalize_vietnam_list)
        if in_list:
            for other_word in list_word:
                if org_word != other_word:  # Avoid comparing the same element
                    other_word = re.sub(r'[.,:_;?/"(){}[]…]', '', other_word)
                    nw = normalize_vietnamese(normalize_word(word))
                    now = normalize_vietnamese(normalize_word(other_word))
                    similar, dist = words_similar(nw, now, threshold = 0.15)
                    if(similar):
                            return begin_punctuation + word + end_punctuation, 2011

            vote_container.append(word)
            weight_container.append(1)
            continue
        for sim in similar_words:
            vote_container.append(sim)
            dist = levenshtein_distance(sim, word)
            weight_container.append(0.54 * (1 - dist/max(len(word), len(sim))))

    if not vote_container:
        return None, 0  # Return None if the list is empty

    if len(vote_container) != len(weight_container):
        raise ValueError("vote_container and weight_container must have the same length.")

        # Aggregate weights for each unique vote
    weight_sum = defaultdict(float)
    for vote, weight in zip(vote_container, weight_container):
        weight_sum[vote] += weight

    # Find the item with the highest total weight
    most_common = max(weight_sum.items(), key=lambda item: item[1])
    most_common_item, total_weight = most_common

    return begin_punctuation + most_common_item + end_punctuation , total_weight


from openpyxl import load_workbook

def read_and_vote(file_path, vietnam_word_list, normalize_vietnam_list ):
    # Load the workbook and select the active sheet
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Iterate through rows starting from the second row
    rowID = 1
    for row in sheet.iter_rows(min_row=2,max_row = 360, min_col=2, max_col=4):  # Columns B, C, D
        print(f"Process on row {rowID}")

        sentence1 = normalize_text(str(row[0].value))  # Column B
        sentence2 = normalize_text(str(row[1].value))   # Column C
        sentence3 = normalize_text(str(row[2].value))  # Column D
        sequences = [sentence1, sentence2, sentence3]
        align_sentences, success = align_multiple_sequences(sequences)
        vote_sentence_splits = []
        if (success == '1'):
            vote_sentence = align_sentences[0]
        else:
            sentences_splits = []
            for align in align_sentences:
                list = align.split(' ')
                sentences_splits.append(list)
            for i in range(len(sentences_splits[0])):
                word1 = sentences_splits[0][i]
                word2 = sentences_splits[1][i]
                word3 = sentences_splits[2][i]

                list_word_to_vote = [word1, word2, word3]

                vote_word = vote(list_word_to_vote,vietnam_word_list, normalize_vietnam_list)

                if(vote_word[0]):
                    vote_sentence_splits.append(vote_word[0])

            vote_sentence = ' '.join(vote_sentence_splits)


        # Write the processed value to column E (fifth column)
        sheet.cell(row=row[0].row, column=5, value=vote_sentence)
        rowID +=1

    # Save changes to the workbook
    workbook.save(file_path)
    print(f"Processed workbook saved: {file_path}")





if __name__ == "__main__":
    # Example usage
    file_path = "ocr_results.xlsx"  # Replace with your file path
    vietnamese_file_path = 'VN_MorphoSyllable_List.txt'
    vietnam_word_list = load_words(vietnamese_file_path)
    normalize_vietnam_list = load_normalize_words(vietnamese_file_path)
    read_and_vote(file_path, vietnam_word_list, normalize_vietnam_list)
    # list_word = ['PHAN', 'PHẦN', 'PHẦN']
    # print(vote(list_word, vietnam_word_list, normalize_vietnam_list))








